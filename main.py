import io
import logging
import openpyxl
import pandas as pd
import requests
import datetime
import time
import schedule
import os
import urllib.parse
import re
from sqlalchemy import create_engine
from dotenv import load_dotenv

# 启动时加载 .env 文件（若存在）。生产环境可直接设置系统环境变量，无需 .env 文件。
load_dotenv()

logger = logging.getLogger('daily_report')

__version__ = '1.1.2'

# ================= 配置区域 =================
# 所有凭据从环境变量读取（优先）或 .env 文件加载（本地开发）。
# 部署时复制 .env.example 为 .env 并填入真实值，或直接设置系统环境变量。

# 1. 数据库配置
DB_CONFIG = {
    'host':     os.environ.get('DB_HOST',     ''),
    'port':     int(os.environ.get('DB_PORT', '3306')),
    'user':     os.environ.get('DB_USER',     ''),
    'password': os.environ.get('DB_PASSWORD', ''),
    'db':       os.environ.get('DB_NAME',     'oil'),
    'charset':  'utf8mb4'
}

# 2. 机器人 Webhook (企业微信)
_webhook_key = os.environ.get('WEBHOOK_KEY', '')
WEBHOOK_URL = os.environ.get(
    'WEBHOOK_URL',
    f"https://qyapi.weixin.qq.com/cgi-bin/webhook/send?key={_webhook_key}"
)

# 4. 配置文件设置
# 方案A：在线表格配置 (填入分享链接)
#   - 腾讯文档：直接填分享链接，程序自动转换为导出链接
#   - 飞书文档：填分享链接 + 配置下方 FEISHU_APP_ID / FEISHU_APP_SECRET
#   - 金山文档/WPS：在文档中「下载 → xlsx」获取直链后填入
CONFIG_EXCEL_URL = os.environ.get('CONFIG_EXCEL_URL', '')

# 飞书开放平台 API 凭据（仅飞书文档需要，其他平台留空即可）
# 获取步骤：
#   1. 访问 https://open.feishu.cn/ → 创建企业自建应用
#   2. 进入「权限管理」→ 搜索并开启 sheets:spreadsheet:readonly
#   3. 进入「版本管理与发布」→ 申请发布（审核通过后生效）
#   4. 在飞书电子表格右上角「分享」→ 添加该应用为协作者（或设为组织内可查看）
#   5. 将 App ID 和 App Secret 填入 .env 文件
FEISHU_APP_ID     = os.environ.get('FEISHU_APP_ID',     '')
FEISHU_APP_SECRET = os.environ.get('FEISHU_APP_SECRET', '')

# 方案B：本地文件配置 (当在线表格未配置或下载失败时使用)
CONFIG_LOCAL_FILE = os.environ.get('CONFIG_LOCAL_FILE', 'device_config.xlsx')

# 5. 飞书告警 Webhook（填入飞书群机器人 Webhook 地址，留空则不推送告警）
# 用于推送两类告警：① 服务异常/终止；② 非致命运行警告（如切换本地配置等）
FEISHU_ALERT_WEBHOOK = os.environ.get('FEISHU_ALERT_WEBHOOK', '')


# =======================================================

def send_feishu_alert(level, title, detail=''):
    """
    推送告警到飞书群机器人。
    level: 'fatal'   → 服务异常终止或人为终止
           'warning' → 非致命问题（如切换本地配置、发送失败等）
    未配置 FEISHU_ALERT_WEBHOOK 时静默跳过。
    """
    if not FEISHU_ALERT_WEBHOOK:
        return
    try:
        prefix   = '[服务停止]' if level == 'fatal' else '[运行警告]'
        now_str  = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        lines    = [f'【智能油库日报机器人】{prefix}', f'时间：{now_str}', f'原因：{title}']
        if detail:
            lines.append(f'详情：{detail}')
        content  = '\n'.join(lines)
        resp = requests.post(
            FEISHU_ALERT_WEBHOOK,
            json={'msg_type': 'text', 'content': {'text': content}},
            timeout=10
        )
        resp.raise_for_status()
        logger.info(f"[飞书告警] 已推送 {prefix}: {title}")
    except Exception as e:
        logger.info(f"[飞书告警] 推送失败（不影响主流程）: {e}")


def setup_logging(log_file):
    """配置日志：同时输出到控制台和文件（每次调用先清除旧 handler，防止多日重复）"""
    logger.handlers.clear()
    logger.setLevel(logging.INFO)
    fmt = logging.Formatter('%(asctime)s  %(message)s', datefmt='%Y-%m-%d %H:%M:%S')
    ch = logging.StreamHandler()
    ch.setFormatter(fmt)
    logger.addHandler(ch)
    fh = logging.FileHandler(log_file, encoding='utf-8')
    fh.setFormatter(fmt)
    logger.addHandler(fh)


def get_db_data():
    """从数据库获取基础设备和库存信息 (使用 SQLAlchemy)"""
    logger.info("正在从数据库读取数据...")

    safe_password = urllib.parse.quote_plus(DB_CONFIG['password'])
    db_url = (
        f"mysql+pymysql://{DB_CONFIG['user']}:{safe_password}"
        f"@{DB_CONFIG['host']}:{DB_CONFIG['port']}/{DB_CONFIG['db']}?charset={DB_CONFIG['charset']}"
    )

    engine = None
    try:
        engine = create_engine(db_url)

        sql = """
        SELECT
            d.device_code,                  -- 设备编号
            d.create_time as install_time,  -- 安装时间
            d.status as device_status,      -- 设备状态：1=停用 2=离线 3=在线
            d.customer_id,                  -- 客户ID（用于排除规则、计量客户筛选）
            c.customer_name,                -- 客户名称
            CONCAT(IFNULL(d.province_name,''), IFNULL(d.city_name,''), IFNULL(d.district_name,'')) as location, -- 安装地点
            ot.oil_model,                   -- 油品型号
            o.avai_ratio,                   -- 当前库存百分比
            o.modify_time                   -- 最后上报时间（用于判断网络在线状态）
        FROM
            t_device d
            -- 关联客户表，获取客户名称
            LEFT JOIN t_customer c ON d.customer_id = c.id
            -- 取每台设备最新一条有效油品配置（status=1 过滤已删除配置，INNER JOIN 取 max(id) 去重）
            LEFT JOIN (
                SELECT ta.* FROM t_oil_type ta
                INNER JOIN ( SELECT device_id, max(id) AS id FROM t_oil_type WHERE status=1 GROUP BY device_id ) tb ON ta.id = tb.id
            ) ot ON d.id = ot.device_id
            -- 取每个油品配置最新一条库存记录（max(id) 去重，不过滤 status，最新插入即为有效值）
            LEFT JOIN (
                SELECT oil_type_id, avai_ratio, modify_time
                FROM t_device_oil
                WHERE id IN (SELECT MAX(id) FROM t_device_oil GROUP BY oil_type_id)
            ) o ON ot.id = o.oil_type_id
        WHERE
            d.del_status = 1        -- 排除已删除设备（del_status: 1=正常 2=删除）
            AND c.id IS NOT NULL    -- 第一道防线：排除无客户关联的设备
            AND c.customer_name IS NOT NULL
            AND c.customer_name != ''
        """
        df = pd.read_sql(sql, engine)
        return df

    except Exception as e:
        logger.info(f"数据库读取失败: {e}")
        send_feishu_alert('warning', '数据库读取失败', str(e))
        return pd.DataFrame()
    finally:
        if engine is not None:
            engine.dispose()



def _parse_xls(xls):
    """从已读取的 Excel 字典中提取四个 sheet"""
    return (
        xls.get('设备配置', pd.DataFrame()),
        xls.get('排除客户设置', pd.DataFrame()),
        xls.get('计量客户设置', pd.DataFrame()),
        xls.get('未录入系统设备', pd.DataFrame()),
    )


def _read_excel_tolerant(filepath):
    """
    读取本地 Excel 文件。
    若 openpyxl 因文件内含不兼容的条件格式规则抛出 ValueError，
    则自动切换到 read_only 流式模式绕过格式解析，仅读取单元格值。
    """
    try:
        return pd.read_excel(filepath, sheet_name=None)
    except Exception as e:
        err = str(e)
        if 'wildcard' not in err and 'numerical' not in err:
            raise  # 非条件格式问题，照常抛出
        logger.info("Excel 含不兼容条件格式，切换兼容模式读取...")
        import openpyxl
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        result = {}
        for sn in wb.sheetnames:
            ws = wb[sn]
            rows = list(ws.iter_rows(values_only=True))
            if rows and rows[0]:
                result[sn] = pd.DataFrame(rows[1:], columns=rows[0])
            else:
                result[sn] = pd.DataFrame()
        wb.close()
        return result


def _fetch_feishu_as_df_dict(url):
    """
    通过飞书开放平台 API 读取电子表格，返回 {sheet名: DataFrame} 字典。
    支持两种链接格式：
      - /sheets/TOKEN：飞书原生电子表格，调用 Sheets API 读取
      - /file/TOKEN  ：云盘文件（上传的 xlsx 或文档移动后链接变为 /file/ 类型），
                       调用 Drive 下载 API 获取二进制后用 openpyxl 解析
    权限要求：
      - /sheets/ 链接需开通 sheets:spreadsheet:readonly
      - /file/   链接需开通 drive:drive:readonly（或 drive:file:readonly）
    """
    m_sheets = re.search(r'/sheets/([A-Za-z0-9_-]+)', url)
    m_file   = re.search(r'/file/([A-Za-z0-9_-]+)',   url)
    if not m_sheets and not m_file:
        raise ValueError("无法从飞书 URL 提取 token，请确认链接格式（应包含 /sheets/ 或 /file/ 路径）")

    # 1. 获取 tenant_access_token（有效期 2 小时，每次使用时实时获取）
    auth_resp = requests.post(
        'https://open.feishu.cn/open-apis/auth/v3/tenant_access_token/internal',
        json={'app_id': FEISHU_APP_ID, 'app_secret': FEISHU_APP_SECRET},
        timeout=10
    )
    auth_resp.raise_for_status()
    auth_data = auth_resp.json()
    if auth_data.get('code') != 0:
        raise RuntimeError(f"飞书认证失败 (code={auth_data.get('code')}): {auth_data.get('msg')}")
    hdrs = {'Authorization': f"Bearer {auth_data['tenant_access_token']}"}

    # ── /file/ 链接：Drive 下载 API → openpyxl 解析 ─────────────────────────
    if m_file and not m_sheets:
        file_token = m_file.group(1)
        logger.info(f"[飞书] 检测到 /file/ 链接，通过 Drive API 下载（token={file_token}）")
        dl_resp = requests.get(
            f'https://open.feishu.cn/open-apis/drive/v1/files/{file_token}/download',
            headers=hdrs, timeout=30
        )
        dl_resp.raise_for_status()
        wb = openpyxl.load_workbook(io.BytesIO(dl_resp.content), data_only=True)
        result = {}
        for sn in wb.sheetnames:
            ws = wb[sn]
            rows = list(ws.iter_rows(values_only=True))
            if rows and rows[0]:
                result[sn] = pd.DataFrame(rows[1:], columns=rows[0])
            else:
                result[sn] = pd.DataFrame()
        wb.close()
        return result

    # ── /sheets/ 链接：Sheets API 读取 ──────────────────────────────────────
    spreadsheet_token = m_sheets.group(1)

    # 2. 读取所有 sheet 元数据（获取 sheetId、title、行列数）
    meta_resp = requests.get(
        f'https://open.feishu.cn/open-apis/sheets/v2/spreadsheets/{spreadsheet_token}/metainfo',
        headers=hdrs, timeout=10
    )
    meta_resp.raise_for_status()
    meta_data = meta_resp.json()
    if meta_data.get('code') != 0:
        raise RuntimeError(f"飞书元数据读取失败: {meta_data.get('msg')}")

    def col_to_letter(n):
        """将列号（1-based）转为 Excel 列字母，如 1→A, 27→AA"""
        s = ''
        while n > 0:
            n, r = divmod(n - 1, 26)
            s = chr(65 + r) + s
        return s

    # 3. 逐 sheet 读取单元格值
    result = {}
    for sheet in meta_data['data']['sheets']:
        sid    = sheet['sheetId']
        title  = sheet['title']
        n_rows = sheet.get('rowCount', 500)
        n_cols = sheet.get('columnCount', 26)
        end_col = col_to_letter(min(n_cols, 52))          # 最多读 AZ（52 列）
        range_str = f'{sid}!A1:{end_col}{n_rows}'

        vr = requests.get(
            f'https://open.feishu.cn/open-apis/sheets/v2/spreadsheets/{spreadsheet_token}/values/{range_str}',
            headers=hdrs, timeout=15
        )
        vr.raise_for_status()
        vdata = vr.json()
        if vdata.get('code') != 0:
            logger.info(f"[飞书] sheet '{title}' 读取失败: {vdata.get('msg')}")
            result[title] = pd.DataFrame()
            continue

        rows = (vdata.get('data') or {}).get('valueRange', {}).get('values') or []
        if not rows:
            result[title] = pd.DataFrame()
            continue

        header = [str(c) if c is not None else '' for c in rows[0]]
        # 裁掉末尾的空列头（飞书按 columnCount 返回整行，右侧空白列会产生多余的 ''）
        while header and header[-1] == '':
            header.pop()
        n_header = len(header)
        data = []
        for row in rows[1:]:
            # 行末空单元格可能被裁剪，补齐到表头长度
            padded = list(row) + [None] * (n_header - len(row))
            data.append(padded[:n_header])
        result[title] = pd.DataFrame(data, columns=header)

    return result


def _check_device_count_diff(online_df_devices, threshold=1):
    """对比在线与本地设备配置的设备编号数量，差异 >= threshold 时推送告警"""
    if not os.path.exists(CONFIG_LOCAL_FILE):
        return
    try:
        local_xls = pd.read_excel(CONFIG_LOCAL_FILE, sheet_name='设备配置', engine='openpyxl')
        col = '设备编号'
        online_codes = set(online_df_devices[col].dropna().astype(str)) if col in online_df_devices.columns else set()
        local_codes  = set(local_xls[col].dropna().astype(str))         if col in local_xls.columns  else set()
        n_online, n_local = len(online_codes), len(local_codes)
        diff = abs(n_online - n_local)
        if diff >= threshold:
            logger.info(f"[内容比对] 在线设备数 {n_online}，本地设备数 {n_local}，差异 {diff} 台，触发告警")
            send_feishu_alert(
                'warning',
                f'在线与本地设备配置差异较大（差异 {diff} 台）',
                f'在线设备数: {n_online}，本地设备数: {n_local}，差异: {diff} 台，请确认配置是否正确。'
            )
        else:
            logger.info(f"[内容比对] 在线设备数 {n_online}，本地设备数 {n_local}，差异 {diff} 台，无需告警")
    except Exception as e:
        logger.info(f"[内容比对] 读取本地配置失败，跳过差异检查: {e}")


def load_config():
    """加载配置：优先在线表格，失败则读取本地Excel"""
    device_config = {}

    df_devices = pd.DataFrame()
    df_settings = pd.DataFrame()
    df_metered_config = pd.DataFrame()
    df_unregistered = pd.DataFrame()

    loaded_source = None

    # 1. 尝试从在线表格读取
    # 支持平台：
    #   - 飞书文档：通过开放平台 API 读取（需配置 FEISHU_APP_ID / FEISHU_APP_SECRET）
    #   - 腾讯文档：自动将分享链接转为导出链接后下载
    #   - 金山文档/WPS：需在文档内「下载 → xlsx」获取直链后填入
    #   - Office Excel 直链：直接下载
    if CONFIG_EXCEL_URL and "http" in CONFIG_EXCEL_URL:
        try:
            logger.info("正在尝试从在线表格加载配置...")

            _url_host = urllib.parse.urlparse(CONFIG_EXCEL_URL).netloc.lower()

            def detect_provider(host):
                if 'docs.qq.com' in host:
                    return '腾讯文档'
                if 'kdocs.cn' in host or 'wps.cn' in host or 'wps.com' in host:
                    return '金山文档/WPS'
                if 'feishu.cn' in host or 'larksuite.com' in host:
                    return '飞书文档'
                return None

            _provider = detect_provider(_url_host)

            # ── 飞书文档：走开放平台 API ──────────────────────────────────
            if _provider == '飞书文档':
                if FEISHU_APP_ID and FEISHU_APP_SECRET:
                    logger.info("检测到飞书文档，通过飞书开放平台 API 读取...")
                    xls = _fetch_feishu_as_df_dict(CONFIG_EXCEL_URL)
                    logger.info(f"[诊断] 在线表格实际包含 Sheet: {sorted(xls.keys())}")
                    _required = {'设备配置', '排除客户设置', '计量客户设置', '未录入系统设备'}
                    _missing = _required - set(xls.keys())
                    if _missing:
                        _s = '、'.join(sorted(_missing))
                        logger.info(f"在线表格缺少以下 Sheet（可能被改名或删除）: {_s}，整体回退本地配置文件")
                        send_feishu_alert('warning', '在线配置与本地配置不同步，已切换本地文件',
                                          f'以下 Sheet 在在线表格中不存在（可能被改名或删除）：{_s}，请检查在线配置文件。')
                    else:
                        df_devices, df_settings, df_metered_config, df_unregistered = _parse_xls(xls)
                        if not df_devices.empty:
                            loaded_source = "在线表格（飞书）"
                            logger.info("飞书在线配置加载成功")
                            _check_device_count_diff(df_devices)
                        else:
                            logger.info("飞书表格无有效数据，将读取本地配置文件")
                            send_feishu_alert('warning', '飞书在线配置表格无有效数据，已切换本地文件')
                else:
                    logger.info("检测到飞书文档，但未配置 FEISHU_APP_ID / FEISHU_APP_SECRET，"
                                "无法通过 API 读取。请参考代码注释完成飞书应用创建后设置环境变量，"
                                "或改用本地文件。")

            # ── 其他平台：HTTP 下载 ───────────────────────────────────────
            else:
                dl_headers = {
                    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36',
                    'Accept': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet, application/octet-stream, */*',
                    'Accept-Language': 'zh-CN,zh;q=0.9',
                }

                def is_xlsx_bytes(content, url_path='', content_disp='', content_type=''):
                    """通过文件头魔术字节或响应头判断内容是否为 Excel 文件"""
                    # XLSX 基于 ZIP 格式，文件头为 PK\x03\x04
                    if content and content[:4] == b"PK\x03\x04":
                        return True
                    if content_disp and ('.xlsx' in content_disp.lower() or '.xls' in content_disp.lower()):
                        return True
                    if url_path and url_path.lower().endswith(('.xlsx', '.xls')):
                        return True
                    if content_type and ('spreadsheet' in content_type or 'vnd.openxmlformats-officedocument' in content_type):
                        return True
                    return False

                def try_request(url):
                    resp = requests.get(url, headers=dl_headers, timeout=20, allow_redirects=True)
                    resp.raise_for_status()
                    return resp

                # 腾讯文档：将分享链接转为导出链接
                fetch_url = CONFIG_EXCEL_URL
                if _provider == '腾讯文档':
                    m = re.search(r"/d/([A-Za-z0-9_-]+)", CONFIG_EXCEL_URL)
                    if m:
                        fetch_url = f"https://docs.qq.com/d/export/{m.group(1)}?format=xlsx"
                        logger.info(f"已将腾讯文档分享链接转换为下载链接: {fetch_url}")

                resp = try_request(fetch_url)
                content_type = resp.headers.get('Content-Type', '')
                content_disp = resp.headers.get('Content-Disposition', '')
                url_path = urllib.parse.urlparse(resp.url).path

                if is_xlsx_bytes(resp.content[:512], url_path, content_disp, content_type):
                    xls = pd.read_excel(io.BytesIO(resp.content), sheet_name=None, engine='openpyxl')
                    _required = {'设备配置', '排除客户设置', '计量客户设置', '未录入系统设备'}
                    _missing = _required - set(xls.keys())
                    if _missing:
                        _s = '、'.join(sorted(_missing))
                        logger.info(f"在线表格缺少以下 Sheet（可能被改名或删除）: {_s}，整体回退本地配置文件")
                        send_feishu_alert('warning', '在线配置与本地配置不同步，已切换本地文件',
                                          f'以下 Sheet 在在线表格中不存在（可能被改名或删除）：{_s}，请检查在线配置文件。')
                    else:
                        df_devices, df_settings, df_metered_config, df_unregistered = _parse_xls(xls)
                        if not df_devices.empty:
                            loaded_source = "在线表格"
                            logger.info("在线配置加载成功")
                            _check_device_count_diff(df_devices)
                        else:
                            logger.info("在线表格无有效数据，将读取本地配置文件")
                            send_feishu_alert('warning', '在线配置表格无有效数据，已切换本地文件')
                else:
                    # 返回的不是 Excel，给出平台相关提示后回退到本地文件
                    if _provider == '腾讯文档':
                        logger.info("腾讯文档导出失败，请检查链接是否有效，或在腾讯文档中点击「文件 → 导出 → Excel」获取导出链接。")
                    elif _provider == '金山文档/WPS':
                        logger.info("金山文档/WPS 分享链接无法自动下载（服务端返回预览页/人机验证）。"
                                    "请在金山文档中点击「下载 → xlsx 格式」，将文件直链填入 CONFIG_EXCEL_URL，或改用本地文件。")
                    else:
                        logger.info("URL返回的是网页而非Excel文件，请填写直接下载链接")

        except Exception as e:
            logger.info(f"在线表格加载失败: {e}")
            send_feishu_alert('warning', '在线配置表格加载失败，已切换本地文件', str(e))

    # 2. 在线数据为空时，回退到本地文件
    if df_devices.empty and os.path.exists(CONFIG_LOCAL_FILE):
        try:
            logger.info(f"正在读取本地配置文件: {CONFIG_LOCAL_FILE}")
            xls = _read_excel_tolerant(CONFIG_LOCAL_FILE)
            df_devices, df_settings, df_metered_config, df_unregistered = _parse_xls(xls)
            if not df_devices.empty:
                loaded_source = "本地文件"
        except Exception as e:
            logger.info(f"本地配置文件读取失败: {e}")
            send_feishu_alert('warning', '本地配置文件读取失败，设备配置将使用默认值', str(e))

    def _row_blank(row):
        """判断一行是否完全空白（所有单元格均为 NaN 或空串），用于跳过预留空行"""
        return all(v is None or (isinstance(v, float) and pd.isna(v)) or str(v).strip() in ('', 'nan')
                   for v in row)

    # 3. 解析设备配置
    _dev_empty_code_rows = []       # 设备编号为空的行号
    _dev_barrel_owner_issues = []   # (行号, 设备编号, 缺失字段描述) — 供后续过滤补录设备后告警
    if not df_devices.empty:
        df_devices.columns = df_devices.columns.str.strip()
        for _row_idx, (_, row) in enumerate(df_devices.iterrows(), start=2):
            if _row_blank(row):
                continue  # 完全空白的预留行，静默跳过
            _code_raw = row.get('设备编号')
            code = '' if (_code_raw is None or (isinstance(_code_raw, float) and pd.isna(_code_raw))) else str(_code_raw).strip()
            if not code or code == 'nan':
                _dev_empty_code_rows.append(_row_idx)
                continue
            barrels = row.get('桶数', 1)
            owner   = row.get('设备归属', '中润')
            device_config[code] = {'barrels': barrels, 'owner': owner}
            # 收集桶数/归属缺失问题（是否为补录设备须在step6后才能判断，此处仅收集）
            _barrel_empty = barrels is None or (isinstance(barrels, float) and pd.isna(barrels))
            _owner_val    = str(owner).strip() if owner is not None else ''
            _owner_empty  = _owner_val in ('', 'nan') or (isinstance(owner, float) and pd.isna(owner))
            if _barrel_empty or _owner_empty:
                _missing = []
                if _barrel_empty: _missing.append('桶数')
                if _owner_empty:  _missing.append('设备归属')
                _dev_barrel_owner_issues.append((_row_idx, code, '、'.join(_missing) + '为空'))
        if _dev_empty_code_rows:
            _detail = '\n'.join(f"  第{r}行 | 原因：设备编号为空" for r in _dev_empty_code_rows)
            logger.info(f"[设备配置] 以下 {len(_dev_empty_code_rows)} 行设备编号为空，已跳过:\n{_detail}")
            send_feishu_alert('warning', f'设备配置校验失败：{len(_dev_empty_code_rows)} 行设备编号为空，请检查"设备配置" Sheet', _detail)

    # 用子串匹配查找列名，容忍列标题带有括号说明文字或尾部空格（如"排除设备编码（...）"）
    def _find_col(columns, keyword):
        for col in columns:
            if col and keyword in str(col):
                return col
        return None

    def _safe_str(val, default=''):
        """将 cell 值安全转为字符串，NaN/None 返回 default"""
        if val is None or (isinstance(val, float) and pd.isna(val)):
            return default
        s = str(val).strip()
        return default if s == 'nan' else s

    def _parse_unreg_df(df):
        """从 '未录入系统设备' DataFrame 解析补录设备列表（列名用子串匹配）
        必填校验：客户名称、设备编号均不可为空，且设备编号在批次内不可重复。
        校验失败的行打印日志并推送飞书告警，不影响其他行的正常解析。
        """
        if df.empty:
            return []
        df = df.copy()
        df.columns = df.columns.astype(str).str.strip()
        logger.info(f"[未录入设备] 读取到 {len(df)} 行，列名: {list(df.columns)}")
        col_code   = _find_col(df.columns, '设备编号') or _find_col(df.columns, '设备编码')
        col_name   = _find_col(df.columns, '客户名称')
        col_time   = _find_col(df.columns, '安装时间') or _find_col(df.columns, '安装日期')
        col_loc    = _find_col(df.columns, '安装地点') or _find_col(df.columns, '安装位置')
        col_barrel = _find_col(df.columns, '桶数')
        col_owner  = _find_col(df.columns, '设备归属')
        logger.info(f"[未录入设备] 列映射 → 设备编号='{col_code}', 客户名称='{col_name}', 安装时间='{col_time}'")
        if not col_code:
            logger.info("[未录入设备] 未找到设备编号列，跳过解析（请确认列标题包含'设备编号'）")
            return []

        devices = []
        failed = []           # 校验失败的条目，格式：(行号, 设备编号或'-', 原因)
        seen_codes = set()    # 批次内已出现的设备编号，用于重复检测

        for row_idx, (_, row) in enumerate(df.iterrows(), start=2):  # start=2：Excel 第1行为表头
            if _row_blank(row):
                continue  # 完全空白的预留行，静默跳过
            code_val = row.get(col_code)
            code = '' if (code_val is None or (isinstance(code_val, float) and pd.isna(code_val))) else str(code_val).strip()
            if not code or code == 'nan':
                reason = '设备编号为空'
                logger.info(f"[未录入设备] 第{row_idx}行录入失败：{reason}")
                failed.append((row_idx, '-', reason))
                continue

            name_val = row.get(col_name) if col_name else None
            name = '' if (name_val is None or (isinstance(name_val, float) and pd.isna(name_val))) else str(name_val).strip()
            if not name or name == 'nan':
                reason = '客户名称为空'
                logger.info(f"[未录入设备] 第{row_idx}行录入失败：设备编号={code}，{reason}")
                failed.append((row_idx, code, reason))
                continue

            if code in seen_codes:
                reason = f'设备编号重复（{code}）'
                logger.info(f"[未录入设备] 第{row_idx}行录入失败：客户名称={name}，{reason}")
                failed.append((row_idx, code, reason))
                continue
            seen_codes.add(code)

            raw_time = row.get(col_time) if col_time else None
            if raw_time is None or (isinstance(raw_time, float) and pd.isna(raw_time)):
                install_time = ''
            elif hasattr(raw_time, 'strftime'):
                install_time = raw_time.strftime('%Y.%m.%d')
            else:
                install_time = _safe_str(raw_time)
            barrel_val = row.get(col_barrel, 1) if col_barrel else 1
            if barrel_val is None or (isinstance(barrel_val, float) and pd.isna(barrel_val)):
                barrel_val = 1
            owner_val = _safe_str(row.get(col_owner) if col_owner else None, '中润') or '中润'
            devices.append({
                '客户名称': name,
                '设备编号': code,
                '安装时间': install_time,
                '安装地点': _safe_str(row.get(col_loc) if col_loc else None),
                '桶数': barrel_val,
                '设备归属': owner_val,
            })

        if failed:
            detail_lines = [f"  第{r}行 | 设备编号={c} | 原因：{rs}" for r, c, rs in failed]
            detail_str = '\n'.join(detail_lines)
            logger.info(f"[未录入设备] 以下 {len(failed)} 条录入失败（校验不通过）:\n{detail_str}")
            send_feishu_alert(
                'warning',
                f'补录设备校验失败，共 {len(failed)} 条，请检查配置文件"未录入系统设备" Sheet',
                detail_str
            )

        return devices

    # 4. 解析排除规则
    # exclusion_rules: [{'customer_id': str, 'device_codes': list}]
    # device_codes 为空 → 排除该客户所有设备；非空 → 只排除列表内的具体设备编号
    exclusion_rules = []
    if not df_settings.empty:
        df_settings.columns = df_settings.columns.str.strip()
        logger.info(f"[配置] 排除客户设置列名: {list(df_settings.columns)}")

        col_cid   = _find_col(df_settings.columns, '排除客户ID')
        col_codes = _find_col(df_settings.columns, '排除设备编码')
        logger.info(f"[配置] 匹配列名 → 排除客户ID='{col_cid}', 排除设备编码='{col_codes}'")

        excl_failed  = []       # 校验失败条目：(行号, 客户ID或'-', 原因)
        seen_excl    = set()    # 批次内重复检测，key=(cid, code) 或 (cid, '')

        for row_idx, (_, row) in enumerate(df_settings.iterrows(), start=2):
            if _row_blank(row):
                continue  # 完全空白的预留行，静默跳过
            # 排除客户ID 必填校验
            cid_raw = row.get(col_cid) if col_cid else None
            if cid_raw is None or (isinstance(cid_raw, float) and pd.isna(cid_raw)):
                excl_failed.append((row_idx, '-', '排除客户ID为空'))
                logger.info(f"[排除规则] 第{row_idx}行解析失败：排除客户ID为空")
                continue
            cid = str(cid_raw).strip()
            if not cid or cid == 'nan':
                excl_failed.append((row_idx, '-', '排除客户ID为空'))
                logger.info(f"[排除规则] 第{row_idx}行解析失败：排除客户ID为空")
                continue
            if '.' in cid:
                cid = cid.split('.')[0]

            # 读取排除设备编码（可为空，空表示排除该客户全部设备）
            codes_raw_val = row.get(col_codes) if col_codes else None
            if codes_raw_val is None or (isinstance(codes_raw_val, float) and pd.isna(codes_raw_val)):
                codes_raw = ''
            else:
                codes_raw = str(codes_raw_val).strip()
            codes = [c.strip() for c in codes_raw.split(',') if c.strip()] if codes_raw else []

            # 重复校验：同一客户ID + 相同设备编码组合不可重复配置
            dedup_key = (cid, tuple(sorted(codes)))
            if dedup_key in seen_excl:
                reason = f'排除规则重复（客户ID={cid}，设备编码={codes_raw or "（全部）"}）'
                excl_failed.append((row_idx, cid, reason))
                logger.info(f"[排除规则] 第{row_idx}行解析失败：{reason}")
                continue
            seen_excl.add(dedup_key)

            logger.info(f"[配置] 排除规则: customer_id={cid}, 设备编码='{codes_raw}' → {codes or '（该客户全部设备）'}")
            exclusion_rules.append({'customer_id': cid, 'device_codes': codes})

        if excl_failed:
            detail_lines = [f"  第{r}行 | 客户ID={c} | 原因：{rs}" for r, c, rs in excl_failed]
            detail_str = '\n'.join(detail_lines)
            logger.info(f"[排除规则] 以下 {len(excl_failed)} 条解析失败:\n{detail_str}")
            send_feishu_alert(
                'warning',
                f'排除客户配置校验失败，共 {len(excl_failed)} 条，请检查配置文件"排除客户设置" Sheet',
                detail_str
            )

    # 5. 解析计量客户ID集合
    metered_customer_ids = set()
    if not df_metered_config.empty:
        df_metered_config.columns = df_metered_config.columns.str.strip()
        if '计量客户ID' in df_metered_config.columns:
            metered_failed = []   # 校验失败条目：(行号, 原因)
            seen_cids = set()     # 批次内重复检测

            for row_idx, (_, row) in enumerate(df_metered_config.iterrows(), start=2):
                if _row_blank(row):
                    continue  # 完全空白的预留行，静默跳过
                # 空值校验
                raw = row.get('计量客户ID')
                if raw is None or (isinstance(raw, float) and pd.isna(raw)):
                    metered_failed.append((row_idx, '-', '计量客户ID为空'))
                    logger.info(f"[计量客户] 第{row_idx}行解析失败：计量客户ID为空")
                    continue
                cid = str(raw).strip()
                if not cid or cid == 'nan':
                    metered_failed.append((row_idx, '-', '计量客户ID为空'))
                    logger.info(f"[计量客户] 第{row_idx}行解析失败：计量客户ID为空")
                    continue
                if '.' in cid:
                    cid = cid.split('.')[0]
                # 重复校验
                if cid in seen_cids:
                    metered_failed.append((row_idx, cid, f'计量客户ID重复（{cid}）'))
                    logger.info(f"[计量客户] 第{row_idx}行解析失败：计量客户ID重复（{cid}）")
                    continue
                seen_cids.add(cid)
                metered_customer_ids.add(cid)

            if metered_failed:
                detail_lines = [f"  第{r}行 | 客户ID={c} | 原因：{rs}" for r, c, rs in metered_failed]
                detail_str = '\n'.join(detail_lines)
                logger.info(f"[计量客户] 以下 {len(metered_failed)} 条解析失败:\n{detail_str}")
                send_feishu_alert(
                    'warning',
                    f'计量客户配置校验失败，共 {len(metered_failed)} 条，请检查配置文件"计量客户设置" Sheet',
                    detail_str
                )

            logger.info(f"[计量客户] 解析完成，有效客户ID {len(metered_customer_ids)} 个: {sorted(metered_customer_ids)}")
        else:
            logger.info("[计量客户] 未找到'计量客户ID'列，跳过解析（请确认列标题正确）")

    # 6. 解析未录入系统设备（补录设备列表）
    supplemental_devices = _parse_unreg_df(df_unregistered)
    if not df_unregistered.empty:
        logger.info(f"未录入系统设备配置加载完成，共 {len(supplemental_devices)} 条")

    # 6a. 桶数/归属校验：补录设备允许为空，其余设备不允许
    _supp_code_set = {dev['设备编号'] for dev in supplemental_devices}
    _non_supp_barrel_owner = [(r, c, rs) for r, c, rs in _dev_barrel_owner_issues if c not in _supp_code_set]
    if _non_supp_barrel_owner:
        _detail_lines = [f"  第{r}行 | 设备编号={c} | {rs}" for r, c, rs in _non_supp_barrel_owner]
        _detail = '\n'.join(_detail_lines)
        logger.info(f"[设备配置] 以下 {len(_non_supp_barrel_owner)} 条非补录设备桶数或设备归属为空:\n{_detail}")
        send_feishu_alert('warning', f'设备配置校验失败：{len(_non_supp_barrel_owner)} 条非补录设备桶数或设备归属为空，请检查"设备配置" Sheet', _detail)

    if loaded_source:
        logger.info(f"配置加载完成 (来源: {loaded_source}，排除规则 {len(exclusion_rules)} 条，计量客户 {len(metered_customer_ids)} 个，补录设备 {len(supplemental_devices)} 台)")
    else:
        logger.info("未加载到有效配置，将使用默认设置")


    return device_config, exclusion_rules, metered_customer_ids, supplemental_devices


def process_data(df):
    """核心逻辑处理"""
    if df.empty:
        return pd.DataFrame(), pd.DataFrame()

    # 1. 加载配置
    device_config, exclusion_rules, metered_customer_ids, supplemental_devices = load_config()

    total_raw = len(df)

    # 2. 过滤掉无客户关联的设备（双保险：SQL 层已过滤，此处兜底应对类型转换边界情况）
    # MySQL varchar NULL → Python None → str(None) = 'None'；numpy NaN → str = 'nan'，统一替换为空串后过滤
    if 'customer_name' in df.columns:
        df['customer_name'] = df['customer_name'].astype(str).str.strip().replace({'nan': '', 'None': ''})
        removed_no_customer = df[df['customer_name'] == '']['device_code'].tolist()
        if removed_no_customer:
            logger.info(f"[诊断] 空客户名被过滤的设备编号: {removed_no_customer}")
        df = df[df['customer_name'] != ''].copy()
    logger.info(f"过滤空客户名后剩余: {len(df)} 台（原始: {total_raw} 台）")

    # 2a. 提前标准化 customer_id（去除 float .0 后缀），后续排除规则和计量客户筛选都依赖此字段
    if 'customer_id' in df.columns:
        df['customer_id'] = df['customer_id'].astype(str).str.strip()
        df['customer_id'] = df['customer_id'].apply(lambda x: x.split('.')[0] if '.' in x else x)

    # 2b. 设备配置 — DB匹配校验 & 总数核对
    # 在排除规则执行前完成，此时 df 代表数据库中全部有效设备（正常+排除+计量）
    # _db_codes 提至块外，供 Step 11 补录设备"已入库"判断使用（含被排除设备）
    _db_codes = set(df['device_code'].astype(str).str.strip())
    if device_config:
        _cfg_codes   = set(device_config.keys())
        _supp_codes  = {dev['设备编号'] for dev in supplemental_devices}
        _supp_not_db = _supp_codes - _db_codes   # 真正不在DB的补录设备

        # C. 非补录设备编码DB匹配校验
        _not_in_db = (_cfg_codes - _supp_codes) - _db_codes
        if _not_in_db:
            _detail = '\n'.join(f"  设备编号={c}" for c in sorted(_not_in_db))
            logger.info(f"[设备配置] 以下 {len(_not_in_db)} 个非补录设备编号在数据库中无匹配，配置可能已失效:\n{_detail}")
            send_feishu_alert(
                'warning',
                f'设备配置中 {len(_not_in_db)} 个设备编号在数据库中无匹配，请确认编号是否正确',
                _detail
            )

        # D. 总数核对：设备配置总数 = 非排除数据库设备数 + 补录不在DB设备数
        # 先根据排除规则计算被命中的DB设备，排除设备无需录入配置表，不计入预期总数
        _excl_db_codes: set = set()
        if exclusion_rules and 'customer_id' in df.columns:
            _df_cid = df['customer_id'].astype(str)
            _df_dcode = df['device_code'].astype(str)
            for _rule in exclusion_rules:
                _rcid  = str(_rule['customer_id'])
                _rcodes = _rule.get('device_codes', [])
                _mask_cid = _df_cid == _rcid
                if _rcodes:
                    _excl_db_codes.update(_df_dcode[_mask_cid & _df_dcode.isin(_rcodes)])
                else:
                    _excl_db_codes.update(_df_dcode[_mask_cid])

        _non_excl_db = _db_codes - _excl_db_codes
        _expected = len(_non_excl_db) + len(_supp_not_db)
        _actual   = len(_cfg_codes)
        if _actual != _expected:
            # 非排除DB设备中未录入配置的设备（这些设备才需要在配置表中）
            _db_not_in_cfg = _non_excl_db - _cfg_codes
            _lines = [
                f"设备配置总数: {_actual} 台，预期: {_expected} 台",
                f"  数据库设备(全量): {len(_db_codes)} 台",
                f"  其中排除规则命中: {len(_excl_db_codes)} 台（不计入预期总数）",
                f"  非排除数据库设备: {len(_non_excl_db)} 台",
                f"  补录设备(不在DB): {len(_supp_not_db)} 台",
            ]
            if _not_in_db:
                _lines.append(f"  配置中无效编号(非补录且不在DB，共 {len(_not_in_db)} 台):")
                _lines.extend(f"    - {c}" for c in sorted(_not_in_db))
            if _db_not_in_cfg:
                _lines.append(f"  非排除DB设备未录入配置(将用默认桶数/归属，共 {len(_db_not_in_cfg)} 台):")
                _lines.extend(f"    - {c}" for c in sorted(_db_not_in_cfg))
            _detail = '\n'.join(_lines)
            logger.info(f"[设备配置] 总数核对不符:\n{_detail}")
            send_feishu_alert(
                'warning',
                f'设备配置总数核对不符：配置 {_actual} 台，预期 {_expected} 台（非排除DB {len(_non_excl_db)} + 补录 {len(_supp_not_db)}）',
                _detail
            )

    # 3. 按排除规则过滤（基于客户ID，精准排除）
    if exclusion_rules and 'customer_id' in df.columns:
        mask = pd.Series(False, index=df.index)
        # 标准化 device_code，防止数据库值含首尾空格导致 isin 匹配失败
        device_code_stripped = df['device_code'].astype(str).str.strip()
        unmatched_rules = []   # 未命中数据库的规则
        for rule in exclusion_rules:
            cid, codes = rule['customer_id'], rule['device_codes']
            if codes:
                # 仅排除该客户下指定的设备编号（旧设备替换场景）
                rule_mask = (df['customer_id'] == cid) & device_code_stripped.isin(codes)
                mask |= rule_mask
                # 按每个设备编码逐一检测是否命中
                unmatched_codes = [c for c in codes if not ((df['customer_id'] == cid) & (device_code_stripped == c)).any()]
                if unmatched_codes:
                    unmatched_rules.append(f"客户ID={cid}，设备编码={','.join(unmatched_codes)} 在数据库中无匹配")
            else:
                # 排除该客户的全部设备
                rule_mask = (df['customer_id'] == cid)
                mask |= rule_mask
                if not rule_mask.any():
                    unmatched_rules.append(f"客户ID={cid} 在数据库中无匹配设备")
        removed_by_rules = df[mask][['customer_id', 'customer_name', 'device_code']].values.tolist()
        if removed_by_rules:
            logger.info(f"[诊断] 排除规则命中的设备（客户ID | 客户名称 | 设备编号）:")
            for item in removed_by_rules:
                logger.info(f"  -> customer_id={item[0]}  {item[1]}  {item[2]}")
        if unmatched_rules:
            detail_str = '\n'.join(f"  {r}" for r in unmatched_rules)
            logger.info(f"[排除规则] 以下规则在数据库中无匹配，排除未生效:\n{detail_str}")
            send_feishu_alert(
                'warning',
                f'排除客户规则在数据库中无匹配，共 {len(unmatched_rules)} 条，请确认客户ID或设备编码是否正确',
                detail_str
            )
        df = df[~mask].copy()
    logger.info(f"排除规则过滤后剩余: {len(df)} 台")

    # 4. 排序（按安装时间升序）
    df['install_time'] = pd.to_datetime(df['install_time'])
    df = df.sort_values(by='install_time', ascending=True)
    df['install_time'] = df['install_time'].dt.strftime('%Y.%m.%d').fillna('')

    # 5. 补全配置字段（桶数、设备归属）
    df['桶数'] = df['device_code'].apply(lambda x: device_config.get(x, {}).get('barrels', 1))
    df['设备归属'] = df['device_code'].apply(lambda x: device_config.get(x, {}).get('owner', '中润'))

    # 6. 网络在线状态判断（直接使用数据库字段）
    now = datetime.datetime.now()

    def check_sync(row):
        # 设备已停用
        if row.get('device_status') == 1:
            return "disabled"
        m_time = row.get('modify_time')
        # 从未上报数据（设备从未激活）
        if pd.isnull(m_time):
            return "never_synced"
        try:
            elapsed_h = (now - m_time).total_seconds() / 3600
            if elapsed_h > 72:
                return "offline"        # 长期失联（>72h），需处理
            if elapsed_h > 24:
                return "offline_warn"   # 短期离线（24~72h），预警
        except Exception:
            return "offline"
        return "online"

    df['网络同步'] = df.apply(check_sync, axis=1)

    # 7. 数据清洗
    df['avai_ratio'] = df['avai_ratio'].fillna('无数据')
    df['oil_model'] = df['oil_model'].fillna('未设置')

    # 7b. 安装地点去除省份前缀，节省列宽（主表和计量客户表均生效）
    if 'location' in df.columns:
        def _strip_province(loc):
            if not loc or not isinstance(loc, str) or str(loc).strip() in ('', 'nan', 'None'):
                return loc
            import re
            s = re.sub(r'^[\u4e00-\u9fa5]{2,4}省', '', str(loc).strip())
            s = re.sub(r'^[\u4e00-\u9fa5]{2,8}自治区', '', s)
            # 直辖市：province_name == city_name，去除重复前缀（如"重庆市重庆市长寿区"→"重庆市长寿区"）
            s = re.sub(r'^([\u4e00-\u9fa5]{2,3}市)\1', r'\1', s)
            return s
        df['location'] = df['location'].apply(_strip_province)

    # 8. 在列重命名前，按 customer_id 筛出计量客户子集
    if metered_customer_ids and 'customer_id' in df.columns:
        df_metered = df[df['customer_id'].isin(metered_customer_ids)].copy()
        # 校验：计量客户ID未命中时，区分「被排除规则过滤」与「数据库真正无记录」两种情形
        # 避免误导维护人员：前者应检查排除规则配置，后者才需要核对数据库
        matched_cids = set(df_metered['customer_id'].unique())
        unmatched = metered_customer_ids - matched_cids
        if unmatched:
            # 全量排除（device_codes 为空）的客户ID集合
            _excl_all_cids = {rule['customer_id'] for rule in exclusion_rules if not rule.get('device_codes')}
            filtered_by_excl = unmatched & _excl_all_cids   # 被排除规则整体过滤
            truly_unmatched  = unmatched - _excl_all_cids   # 数据库真正无匹配
            if filtered_by_excl:
                detail_str = '、'.join(sorted(filtered_by_excl))
                logger.info(f"[计量客户] 以下客户ID被排除规则整体过滤，计量Sheet中无对应数据，"
                            f"请检查排除客户设置是否与计量客户设置冲突: {detail_str}")
                send_feishu_alert(
                    'warning',
                    f'计量客户与排除规则冲突，共 {len(filtered_by_excl)} 个客户ID，计量Sheet无数据',
                    f'以下计量客户ID被排除规则整体过滤，请检查配置冲突: {detail_str}'
                )
            if truly_unmatched:
                detail_str = '、'.join(sorted(truly_unmatched))
                logger.info(f"[计量客户] 以下客户ID在数据库中无匹配设备，请确认ID是否正确: {detail_str}")
                send_feishu_alert(
                    'warning',
                    f'计量客户ID在数据库中无匹配设备，共 {len(truly_unmatched)} 个，请检查配置',
                    f'未匹配的客户ID: {detail_str}'
                )
    else:
        df_metered = pd.DataFrame()

    # 9. 生成序号
    df.reset_index(drop=True, inplace=True)
    df.insert(0, '序号', pd.Series(range(1, 1 + len(df))))

    # 10. 列重命名与筛选
    rename_map = {
        'customer_name': '客户名称',
        'device_code': '设备编号',
        'oil_model': '油品型号',
        'avai_ratio': '库存(%)',
        'install_time': '安装时间',
        'location': '安装地点'
    }
    target_cols = ['序号', '客户名称', '设备编号', '油品型号', '库存(%)', '桶数', '设备归属', '网络同步',
                   '安装时间', '安装地点']

    df = df.rename(columns=rename_map)
    df_out = df[[c for c in target_cols if c in df.columns]]

    # 11. 补录配置文件中"未录入系统设备"的设备
    if supplemental_devices:
        supp_rows = []
        supp_found_in_db = []   # 补录设备编号已在DB中，需提醒维护人员清理
        for dev in supplemental_devices:
            code = dev.get('设备编号', '')
            # 以全量 _db_codes 判断"已入库"，而非 existing_codes（报表中已过滤排除设备）
            # 修复：设备在DB但被排除规则过滤时，existing_codes 不含该设备，
            # 原逻辑会将其重新追加为补录设备（disabled），导致已排除设备出现在报表末尾
            if code in _db_codes:
                logger.info(f'[未录入设备] 设备 {code} 已在数据库中，已按正常设备处理，请从配置文件"未录入系统设备" Sheet 中删除该记录')
                supp_found_in_db.append(code)
                continue
            supp_rows.append({
                '客户名称': dev.get('客户名称', ''),
                '设备编号': code,
                '油品型号': '设备待安装调试且未录入系统',
                '库存(%)': '无数据',
                '桶数': dev.get('桶数', 1),
                '设备归属': dev.get('设备归属', '中润'),
                '网络同步': 'disabled',
                '安装时间': dev.get('安装时间', ''),
                '安装地点': dev.get('安装地点', ''),
            })
            logger.info(f"[未录入设备] 补录成功：客户名称={dev.get('客户名称', '')}，设备编号={code}，"
                        f"安装时间={dev.get('安装时间', '') or '未填写'}，安装地点={dev.get('安装地点', '') or '未填写'}")
        if supp_found_in_db:
            _detail = '、'.join(supp_found_in_db)
            send_feishu_alert(
                'warning',
                f'补录设备已入库，请及时清理配置，共 {len(supp_found_in_db)} 台',
                f'以下设备编号已在数据库中正常存在，已按正常设备处理，请从"未录入系统设备" Sheet 中删除：{_detail}'
            )
        if supp_rows:
            df_supp = pd.DataFrame(supp_rows)
            df_out_no_seq = df_out.drop(columns=['序号'])
            supp_cols = [c for c in df_out_no_seq.columns if c in df_supp.columns]
            df_out_no_seq = pd.concat([df_out_no_seq, df_supp[supp_cols]], ignore_index=True)
            df_out_no_seq.insert(0, '序号', pd.Series(range(1, 1 + len(df_out_no_seq))))
            df_out = df_out_no_seq
            logger.info(f"已补录 {len(supp_rows)} 台未入库设备，报表共 {len(df_out)} 台")

    # 对计量客户子集做同样的重命名与列筛选（序号独立编排）
    if not df_metered.empty:
        df_metered.reset_index(drop=True, inplace=True)
        df_metered.insert(0, '序号', pd.Series(range(1, 1 + len(df_metered))))
        df_metered = df_metered.rename(columns=rename_map)
        df_metered = df_metered[[c for c in target_cols if c in df_metered.columns]]

    return df_out, df_metered


def generate_excel_with_format(df, filename, df_metered=None):
    """生成Excel：主表 + 可选计量客户单列 sheet"""
    writer = pd.ExcelWriter(filename, engine='xlsxwriter')

    main_sheet = '安卓设备日统计表'
    metered_sheet = '计量客户单列'
    today_str = datetime.datetime.now().strftime('%Y年%m月%d日')

    # 数据从第2行写起（第1行留给标题行）
    df.to_excel(writer, index=False, sheet_name=main_sheet, startrow=1)
    if df_metered is not None and not df_metered.empty:
        df_metered.to_excel(writer, index=False, sheet_name=metered_sheet, startrow=1)

    workbook = writer.book

    # === 颜色 ===
    C_HEADER_BG = '#2E75B6'   # 中蓝（表头背景）
    C_TITLE_BG  = '#1F4E79'   # 深蓝（标题背景）
    C_BORDER    = '#B8D4E8'   # 浅蓝边框
    C_ROW_EVEN  = '#F0F5FA'   # 偶数行：冷灰蓝（柔和、专业）
    FONT        = '微软雅黑'

    # 居中列：数字/状态类；其余文本列左对齐
    CENTER_COLS = {'序号', '库存(%)', '桶数', '设备归属', '网络同步', '安装时间'}

    C_OUTER_B = '#2E75B6'   # 外框线颜色（与表头一致，视觉统一）

    # === 格式对象 ===
    title_fmt = workbook.add_format({
        'bold': True, 'font_name': FONT, 'font_size': 16,
        'align': 'center', 'valign': 'vcenter',
        'bg_color': C_TITLE_BG, 'font_color': '#FFFFFF',
        'top': 2, 'top_color': C_TITLE_BG,       # 顶部粗线与背景同色（融入）
        'bottom': 2, 'bottom_color': C_HEADER_BG, # 底部粗线作为与表头的分隔
        'left': 2, 'left_color': C_OUTER_B,
        'right': 2, 'right_color': C_OUTER_B,
    })
    header_fmt = workbook.add_format({
        'bold': True, 'font_name': FONT, 'font_size': 11,
        'align': 'center', 'valign': 'vcenter',
        'bg_color': C_HEADER_BG, 'font_color': '#FFFFFF',
        'border': 2, 'border_color': '#1F4E79',   # 四边粗线，强化表头分隔
    })

    # 数据单元格按需创建格式：外框粗线(2)，内框细线(1)，奇偶底色，居中/左对齐，特殊色
    # 使用缓存避免重复创建，最终格式对象数 ≤ 50 个
    _fmt_cache = {}

    def _make_fmt(bg, align, font_color, top_b, bottom_b, left_b, right_b):
        key = (bg, align, font_color, top_b, bottom_b, left_b, right_b)
        if key not in _fmt_cache:
            def _bc(w): return C_OUTER_B if w == 2 else C_BORDER
            props = {
                'font_name': FONT, 'font_size': 10, 'valign': 'vcenter',
                'bg_color': bg, 'align': align,
                'top':    top_b,    'top_color':    _bc(top_b),
                'bottom': bottom_b, 'bottom_color': _bc(bottom_b),
                'left':   left_b,   'left_color':   _bc(left_b),
                'right':  right_b,  'right_color':  _bc(right_b),
            }
            if font_color:
                props['font_color'] = font_color
            _fmt_cache[key] = workbook.add_format(props)
        return _fmt_cache[key]

    def _inv_colors(val):
        """根据库存值返回 (bg_color, font_color) 元组"""
        if val == '无数据' or val is None or (isinstance(val, float) and pd.isna(val)):
            return '#FFCC99', '#333333'
        try:
            v = float(val)
            if v > 30: return '#C6EFCE', '#006100'
            if v > 5:  return '#FFEB9C', '#9C5700'
            return '#FFC7CE', '#9C0006'
        except (ValueError, TypeError):
            return '#FFCC99', '#333333'

    _SYNC_COLOR_MAP = {
        'online':       ('#C6EFCE', '#006100'),
        'offline_warn': ('#FFD966', '#7D4C00'),
        'offline':      ('#FFC7CE', '#9C0006'),
        'never_synced': ('#E2CFEE', '#7030A0'),
        'disabled':     ('#D9D9D9', '#666666'),
    }

    def _sync_colors(val):
        """根据网络状态值返回 (bg_color, font_color) 元组"""
        return _SYNC_COLOR_MAP.get(val, ('#D9D9D9', '#666666'))

    # 各列宽度（列名 → 字符宽度）
    # 网络同步 -2 英文，客户名称/油品型号各 +4 汉字（+8），设备编号 +3 英文
    col_widths = {
        '序号': 4, '客户名称': 26, '设备编号': 18,
        '油品型号': 22, '库存(%)': 7, '桶数': 5,
        '设备归属': 7, '网络同步': 11, '安装时间': 10, '安装地点': 15,
    }

    def apply_sheet_format(ws, sheet_df, sheet_title):
        """对单个 worksheet 应用标题行 + 表头 + 交替行底色 + 条件格式 + 保护"""
        num_cols    = len(sheet_df.columns)
        last_col    = num_cols - 1
        last_data_row = 1 + len(sheet_df)   # 0-indexed：row0=标题 row1=表头 row2..=数据

        # ── 标题行（row 0）──
        ws.set_row(0, 45)
        ws.merge_range(0, 0, 0, last_col, sheet_title, title_fmt)

        # ── 表头行（row 1，覆盖 pandas 写入的默认格式）──
        ws.set_row(1, 26)
        for col_num, col_name in enumerate(sheet_df.columns.values):
            ws.write(1, col_num, col_name, header_fmt)
            if col_name == '网络同步':
                ws.write_comment(1, col_num,
                    '【网络同步状态说明】\n'
                    'online — 24h 内正常上报（绿）\n'
                    'offline_warn — 24~72h 未上报（橙）\n'
                    'offline — 超过 72h 未上报（红）\n'
                    'never_synced — 从未上报，设备未激活（紫）\n'
                    'disabled — 已停用 / 待安装调试（灰）',
                    {'x_scale': 2.5, 'y_scale': 3.5, 'font_size': 10}
                )

        # ── 列宽 ──
        for col_num, col_name in enumerate(sheet_df.columns.values):
            ws.set_column(col_num, col_num, col_widths.get(col_name, 12))

        # ── 数据行：按边缘位置决定粗/细边框，直接写入格式（兼容所有阅读器）──
        total_rows = len(sheet_df)
        for row_num in range(total_rows):
            excel_row = row_num + 2
            ws.set_row(excel_row, 20)
            is_even = (excel_row % 2 == 0)
            t_b = 2 if row_num == 0             else 1  # 首行顶边粗
            b_b = 2 if row_num == total_rows - 1 else 1  # 末行底边粗
            bg_row = C_ROW_EVEN if is_even else '#FFFFFF'
            for col_num, col_name in enumerate(sheet_df.columns.values):
                val   = sheet_df.iloc[row_num, col_num]
                l_b   = 2 if col_num == 0         else 1  # 首列左边粗
                r_b   = 2 if col_num == last_col   else 1  # 末列右边粗
                align = 'center' if col_name in CENTER_COLS else 'left'
                if col_name == '库存(%)':
                    bg, fc = _inv_colors(val)
                elif col_name == '网络同步':
                    bg, fc = _sync_colors(str(val))
                else:
                    bg, fc = bg_row, None
                ws.write(excel_row, col_num, val,
                         _make_fmt(bg, align, fc, t_b, b_b, l_b, r_b))

        # ── 冻结标题+表头两行 ──
        ws.freeze_panes(2, 0)

        # ── 自动筛选（作用在表头行） ──
        ws.autofilter(1, 0, last_data_row, last_col)

        # ── 打印设置：最小页边距 + 竖横向自适应 + 页脚页码 + 重复表头 ──
        ws.print_area(0, 0, last_data_row, last_col)      # 显式声明打印区域，确保末列不被截断
        ws.set_paper(9)                                    # A4
        ws.fit_to_pages(1, 0)                             # 宽度1页，行数不限（竖/横向通用）
        ws.set_margins(left=0.25, right=0.25, top=0.5, bottom=0.5)  # 最小侧边距
        ws.set_footer('&C第 &P 页 / 共 &N 页')            # 居中页码
        ws.repeat_rows(0, 1)                              # 每页重复标题行+表头行

    apply_sheet_format(
        writer.sheets[main_sheet], df,
        f'Android版智能油库油量统计表（{today_str}）'
    )
    if df_metered is not None and not df_metered.empty:
        apply_sheet_format(
            writer.sheets[metered_sheet], df_metered,
            f'Android版智能油库油量统计表 · 计量客户（{today_str}）'
        )

    writer.close()
    return filename



def send_to_robot(filename):
    """发送文件"""
    if not os.path.exists(filename) or "YOUR_WEBHOOK" in WEBHOOK_URL:
        logger.info("未配置Webhook或文件不存在，跳过发送。")
        return

    try:
        # 用标准库解析 key，避免因 URL 参数顺序变化导致提取错误
        parsed = urllib.parse.urlparse(WEBHOOK_URL)
        key = urllib.parse.parse_qs(parsed.query).get('key', [''])[0]
        if not key:
            logger.info("Webhook URL 中未找到 key 参数，跳过发送。")
            return
        upload_url = f"https://qyapi.weixin.qq.com/cgi-bin/webhook/upload_media?key={key}&type=file"

        with open(filename, 'rb') as f:
            files = {'file': (os.path.basename(filename), f,
                              'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
            resp = requests.post(upload_url, files=files, timeout=30)
            media_id = resp.json().get('media_id')

        if media_id:
            msg = {"msgtype": "file", "file": {"media_id": media_id}}
            requests.post(WEBHOOK_URL, json=msg, timeout=15)
            logger.info("报表发送成功")
        else:
            logger.info(f"上传失败: {resp.text}")
            send_feishu_alert('warning', '报表上传失败', resp.text)
    except Exception as e:
        logger.info(f"发送异常: {e}")
        send_feishu_alert('warning', '报表发送异常', str(e))


def clean_old_files(keep=10):
    """清理旧的报表和日志文件，按修改时间排序，超出保留数量时先删最旧的。"""
    logger.info(f"开始清理旧报表及日志文件（每类最多保留 {keep} 份）...")
    try:
        cwd = os.path.abspath('.')
        reports, logs = [], []
        for f in os.listdir(cwd):
            path = os.path.join(cwd, f)
            if not os.path.isfile(path):
                continue
            mtime = os.path.getmtime(path)
            if f.endswith('.xlsx') and '智能油库油量统计表' in f:
                reports.append((mtime, path))
            elif f.endswith('.log') and '智能油库运行日志' in f:
                logs.append((mtime, path))

        deleted = []
        for group in (reports, logs):
            group.sort()                    # 升序：最旧在前
            to_delete = group[:-keep] if len(group) > keep else []
            for _, path in to_delete:
                try:
                    os.remove(path)
                    deleted.append(os.path.basename(path))
                except Exception as e:
                    logger.warning(f"删除文件 {os.path.basename(path)} 失败: {e}")

        if deleted:
            logger.info(f"清理完成，已删除 {len(deleted)} 个旧文件: {', '.join(deleted)}")
        else:
            logger.info("清理完成，无需删除文件")
    except Exception as e:
        logger.warning(f"清理文件过程出错: {e}")


def daily_task():
    """主任务"""
    today_str = datetime.datetime.now().strftime('%Y年%m月%d日')
    log_file = f"{today_str}智能油库运行日志.log"

    # 先初始化日志，再清理旧文件（清理结果同步写入日志）
    setup_logging(log_file)
    clean_old_files()

    logger.info(f"智能油库日报机器人 v{__version__} 开始执行，日志文件: {log_file}")

    df_db = get_db_data()

    if df_db.empty:
        logger.info("无数据库数据")
        return

    df_final, df_metered = process_data(df_db)

    filename = f"{today_str}智能油库油量统计表.xlsx"

    try:
        generate_excel_with_format(df_final, filename, df_metered)
    except Exception as e:
        logger.info(f"Excel 生成失败，终止本次任务: {e}")
        send_feishu_alert('warning', 'Excel 报表生成失败，本次任务已终止', str(e))
        return

    send_to_robot(filename)
    logger.info(f"处理完成: {filename}")


if __name__ == "__main__":
    print("=== 机器人运行中 ===")

    # 启动时立即执行一次
    daily_task()  # 仅限程序启动时测试使用（部署前请注释此行，防止双次执行）

    # 注册每日 08:00 定时任务
    # schedule 以"距上次执行是否已满 24 小时"判断是否触发。
    # 若程序在 08:00 之后启动，调度器会认为本次窗口尚未执行过，
    # 并在第一次 run_pending() 时立即触发第二次执行（与启动时的 daily_task() 重复）。
    # 将 last_run 设为当前时间，相当于告知调度器"本次窗口已执行"，从而跳过此次触发，
    # 等到明天 08:00 再正常触发。
    job = schedule.every().day.at("08:00").do(daily_task)
    job.last_run = datetime.datetime.now()

    try:
        while True:
            schedule.run_pending()
            time.sleep(60)
    except KeyboardInterrupt:
        logger.info("程序被用户中断，正在优雅退出...")
        send_feishu_alert('fatal', '程序被人为中断（Ctrl+C 或系统停止信号）')
        logger.info("程序已安全退出")
    except Exception as e:
        logger.info(f"程序异常终止: {e}")
        send_feishu_alert('fatal', f'程序异常终止: {type(e).__name__}', str(e))
        raise
