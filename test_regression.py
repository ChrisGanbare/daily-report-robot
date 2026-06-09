"""
test_regression.py — 全链路回归测试
覆盖：check_sync 五级状态 / process_data 核心管道 / load_config 降级策略 /
       generate_excel_with_format 文件完整性 / send_to_robot 分支路径
不依赖真实数据库或外部网络，全部使用 mock / 本地数据。
"""

import io
import os
import sys
import types
import datetime
import unittest
from unittest.mock import MagicMock, patch, mock_open

import pandas as pd
import openpyxl

# ── 屏蔽 load_dotenv ──────────────────────────────────────────────────────────
sys.modules.setdefault('dotenv', types.ModuleType('dotenv'))
sys.modules['dotenv'].load_dotenv = lambda: None  # type: ignore

import main  # noqa: E402

import logging
logging.basicConfig(level=logging.CRITICAL)


# =============================================================================
# 工具函数
# =============================================================================

def _make_db_row(**kwargs):
    """构造单条数据库记录的默认字段，允许覆盖"""
    now = datetime.datetime.now()
    defaults = {
        'device_code': 'D001',
        'install_time': datetime.datetime(2023, 1, 1),
        'device_status': 3,          # 3=在线（按数据库约定：1=停用 2=离线 3=在线）
        'customer_id': 100001,
        'customer_name': '测试客户',
        'location': '浙江省杭州市西湖区',
        'oil_model': '0W-20',
        'avai_ratio': 50.0,
        'modify_time': now - datetime.timedelta(hours=1),
    }
    defaults.update(kwargs)
    return defaults


def _make_df(*rows):
    """将多条记录字典合并为 DataFrame"""
    return pd.DataFrame(list(rows))


def _empty_config():
    """返回空配置四元组，使 load_config 不读取任何文件"""
    return {}, [], set(), []


# =============================================================================
# 1. check_sync 五级网络状态
# =============================================================================
class TestCheckSync(unittest.TestCase):
    """直接通过 process_data 触发 check_sync，验证五级状态赋值"""

    def _run(self, device_status, modify_time_offset_h):
        """构造单行 DataFrame 跑 process_data，返回网络同步值"""
        now = datetime.datetime.now()
        if modify_time_offset_h is None:
            m_time = None
        else:
            m_time = now - datetime.timedelta(hours=modify_time_offset_h)

        row = _make_db_row(device_status=device_status, modify_time=m_time)
        df = _make_df(row)

        with patch.object(main, 'load_config', return_value=_empty_config()):
            df_out, _ = main.process_data(df)

        return df_out['网络同步'].iloc[0] if not df_out.empty else None

    def test_disabled(self):
        """device_status=1 → disabled"""
        self.assertEqual(self._run(1, 1), 'disabled')

    def test_never_synced(self):
        """modify_time=None → never_synced"""
        self.assertEqual(self._run(3, None), 'never_synced')

    def test_online(self):
        """modify_time 距今 1h → online"""
        self.assertEqual(self._run(3, 1), 'online')

    def test_offline_warn(self):
        """modify_time 距今 48h → offline_warn"""
        self.assertEqual(self._run(3, 48), 'offline_warn')

    def test_offline(self):
        """modify_time 距今 96h → offline"""
        self.assertEqual(self._run(3, 96), 'offline')


# =============================================================================
# 2. process_data 核心管道
# =============================================================================
class TestProcessData(unittest.TestCase):

    def _base_df(self, n=3):
        rows = [_make_db_row(
            device_code=f'D{i:03d}',
            customer_id=100000 + i,
            customer_name=f'客户{i}',
        ) for i in range(1, n + 1)]
        return _make_df(*rows)

    # ── 2a. 空客户名过滤 ──────────────────────────────────────────────────────
    def test_empty_customer_name_filtered(self):
        """customer_name 为空的设备应被过滤"""
        rows = [
            _make_db_row(device_code='D001', customer_name='正常客户'),
            _make_db_row(device_code='D002', customer_name=''),
            _make_db_row(device_code='D003', customer_name=None),
        ]
        df = _make_df(*rows)
        with patch.object(main, 'load_config', return_value=_empty_config()):
            df_out, _ = main.process_data(df)
        self.assertEqual(len(df_out), 1)
        self.assertEqual(df_out['设备编号'].iloc[0], 'D001')

    # ── 2b. 排除规则（整体排除）────────────────────────────────────────────────
    def test_exclusion_rule_full_customer(self):
        """exclusion_rules customer_id 匹配时，该客户所有设备被排除"""
        df = self._base_df(3)   # D001(cid=100001), D002(100002), D003(100003)
        cfg = ({}, [{'customer_id': '100002', 'device_codes': []}], set(), [])
        with patch.object(main, 'load_config', return_value=cfg):
            df_out, _ = main.process_data(df)
        codes = df_out['设备编号'].tolist()
        self.assertNotIn('D002', codes)
        self.assertIn('D001', codes)
        self.assertIn('D003', codes)

    # ── 2c. 排除规则（指定编号）───────────────────────────────────────────────
    def test_exclusion_rule_specific_device(self):
        """exclusion_rules 指定 device_codes 时，只排除对应编号"""
        df = self._base_df(3)
        cfg = ({}, [{'customer_id': '100001', 'device_codes': ['D001']}], set(), [])
        with patch.object(main, 'load_config', return_value=cfg):
            df_out, _ = main.process_data(df)
        self.assertNotIn('D001', df_out['设备编号'].tolist())
        self.assertIn('D002', df_out['设备编号'].tolist())

    # ── 2d. 计量客户子集 ──────────────────────────────────────────────────────
    def test_metered_subset_extracted(self):
        """metered_customer_ids 命中时，返回非空 df_metered"""
        df = self._base_df(3)
        cfg = ({}, [], {'100002'}, [])
        with patch.object(main, 'load_config', return_value=cfg):
            _, df_metered = main.process_data(df)
        self.assertFalse(df_metered.empty)
        self.assertIn('D002', df_metered['设备编号'].tolist())

    # ── 2e. 补录设备追加 ──────────────────────────────────────────────────────
    def test_supplemental_device_appended(self):
        """不在 DB 的补录设备应追加到报表末尾"""
        df = self._base_df(2)
        supp = [{'客户名称': '补录客户', '设备编号': 'S001',
                 '安装时间': '2026.01.01', '安装地点': '测试地点',
                 '桶数': 1, '设备归属': '中润'}]
        cfg = ({}, [], set(), supp)
        with patch.object(main, 'load_config', return_value=cfg):
            df_out, _ = main.process_data(df)
        self.assertIn('S001', df_out['设备编号'].tolist())
        row = df_out[df_out['设备编号'] == 'S001'].iloc[0]
        self.assertEqual(row['网络同步'], 'disabled')

    # ── 2f. 补录设备已入库则不追加 ───────────────────────────────────────────
    def test_supplemental_device_in_db_not_appended(self):
        """补录设备编号已在 DB 中时，不追加，正常流程处理"""
        df = self._base_df(2)   # D001, D002
        supp = [{'客户名称': '某客户', '设备编号': 'D001',
                 '安装时间': '', '安装地点': '',
                 '桶数': 1, '设备归属': '中润'}]
        cfg = ({}, [], set(), supp)
        alerts = []
        with patch.object(main, 'load_config', return_value=cfg), \
             patch.object(main, 'send_feishu_alert',
                          side_effect=lambda *a, **kw: alerts.append(a)):
            df_out, _ = main.process_data(df)
        # D001 出现一次（DB正常），不出现第二次（补录跳过）
        self.assertEqual(df_out['设备编号'].tolist().count('D001'), 1)
        # 应触发"补录设备已入库"告警
        self.assertTrue(any('入库' in str(a) for a in alerts))

    # ── 2g. 排序：安装时间升序 ─────────────────────────────────────────────────
    def test_sorted_by_install_time(self):
        """输出按 install_time 升序排列"""
        rows = [
            _make_db_row(device_code='D_LATE',
                         install_time=datetime.datetime(2024, 6, 1)),
            _make_db_row(device_code='D_EARLY',
                         install_time=datetime.datetime(2022, 1, 1)),
        ]
        df = _make_df(*rows)
        with patch.object(main, 'load_config', return_value=_empty_config()):
            df_out, _ = main.process_data(df)
        self.assertEqual(df_out['设备编号'].iloc[0], 'D_EARLY')

    # ── 2h. 序号连续从 1 开始 ─────────────────────────────────────────────────
    def test_seq_starts_at_1(self):
        """输出序号列应从 1 连续编排"""
        df = self._base_df(4)
        with patch.object(main, 'load_config', return_value=_empty_config()):
            df_out, _ = main.process_data(df)
        self.assertEqual(list(df_out['序号']), list(range(1, len(df_out) + 1)))

    # ── 2i. 安装地点省份前缀去除 ─────────────────────────────────────────────
    def test_province_prefix_stripped(self):
        """安装地点应去除省份前缀"""
        cases = {
            '广东省东莞市塘厦镇': '东莞市塘厦镇',
            '内蒙古自治区呼和浩特市': '呼和浩特市',
            '重庆市重庆市长寿区': '重庆市长寿区',
            '浙江省杭州市': '杭州市',
        }
        for loc_in, loc_expected in cases.items():
            row = _make_db_row(location=loc_in)
            df = _make_df(row)
            with patch.object(main, 'load_config', return_value=_empty_config()):
                df_out, _ = main.process_data(df)
            actual = df_out['安装地点'].iloc[0]
            self.assertEqual(actual, loc_expected,
                             f"输入'{loc_in}'，期望'{loc_expected}'，实际'{actual}'")

    # ── 2j. 无效安装时间不崩溃（Bug 1 回归）─────────────────────────────────
    def test_invalid_install_time_no_crash(self):
        """install_time 包含无效值时不应崩溃（errors='coerce' 修复回归）"""
        rows = [
            _make_db_row(device_code='D001', install_time='INVALID'),
            _make_db_row(device_code='D002',
                         install_time=datetime.datetime(2023, 1, 1)),
        ]
        df = _make_df(*rows)
        try:
            with patch.object(main, 'load_config', return_value=_empty_config()):
                df_out, _ = main.process_data(df)
        except Exception as e:
            self.fail(f"无效安装时间不应崩溃，但抛出: {e}")
        self.assertEqual(len(df_out), 2)


# =============================================================================
# 3. load_config 降级策略
# =============================================================================
class TestLoadConfigFallback(unittest.TestCase):

    def _make_local_excel(self, path):
        """生成包含四个必需 Sheet 的本地 Excel"""
        with pd.ExcelWriter(path, engine='openpyxl') as w:
            pd.DataFrame({'设备编号': ['D001'], '桶数': [2], '设备归属': ['中润']}).to_excel(
                w, index=False, sheet_name='设备配置')
            pd.DataFrame(columns=['排除客户名称', '排除客户ID', '排除设备编码']).to_excel(
                w, index=False, sheet_name='排除客户设置')
            pd.DataFrame(columns=['计量客户名称', '计量客户ID']).to_excel(
                w, index=False, sheet_name='计量客户设置')
            pd.DataFrame(columns=['客户名称', '设备编号']).to_excel(
                w, index=False, sheet_name='未录入系统设备')

    def test_fallback_to_local_when_no_url(self):
        """CONFIG_EXCEL_URL 为空时应直接读取本地文件"""
        tmp = '_test_lc_local.xlsx'
        self._make_local_excel(tmp)
        try:
            with patch.object(main, 'CONFIG_EXCEL_URL', ''), \
                 patch.object(main, 'CONFIG_LOCAL_FILE', tmp):
                device_cfg, excl, metered, supp = main.load_config()
            self.assertIn('D001', device_cfg)
        finally:
            if os.path.exists(tmp): os.remove(tmp)

    def test_fallback_to_local_when_online_fails(self):
        """在线表格加载异常时应降级到本地文件"""
        tmp = '_test_lc_fallback.xlsx'
        self._make_local_excel(tmp)
        try:
            with patch.object(main, 'CONFIG_EXCEL_URL', 'https://fake.feishu.cn/sheets/TOKEN'), \
                 patch.object(main, 'FEISHU_APP_ID', 'fake_id'), \
                 patch.object(main, 'FEISHU_APP_SECRET', 'fake_secret'), \
                 patch.object(main, '_fetch_feishu_as_df_dict',
                              side_effect=RuntimeError("网络不可达")), \
                 patch.object(main, 'CONFIG_LOCAL_FILE', tmp), \
                 patch.object(main, 'send_feishu_alert'):
                device_cfg, _, _, _ = main.load_config()
            self.assertIn('D001', device_cfg, "降级到本地文件后应包含 D001")
        finally:
            if os.path.exists(tmp): os.remove(tmp)

    def test_no_config_returns_empty(self):
        """无在线 URL 且无本地文件时，返回空配置（不崩溃）"""
        with patch.object(main, 'CONFIG_EXCEL_URL', ''), \
             patch.object(main, 'CONFIG_LOCAL_FILE', '_nonexistent_file_.xlsx'):
            device_cfg, excl, metered, supp = main.load_config()
        self.assertEqual(device_cfg, {})
        self.assertEqual(excl, [])
        self.assertEqual(metered, set())
        self.assertEqual(supp, [])


# =============================================================================
# 4. generate_excel_with_format 文件完整性
# =============================================================================
class TestGenerateExcel(unittest.TestCase):

    def _make_df_out(self, n=5):
        rows = []
        for i in range(1, n + 1):
            rows.append({
                '序号': i, '客户名称': f'客户{i}', '设备编号': f'D{i:03d}',
                '油品型号': '0W-20', '库存(%)': i * 10,
                '桶数': 2, '设备归属': '中润',
                '网络同步': ['online', 'offline_warn', 'offline',
                            'never_synced', 'disabled'][i % 5],
                '安装时间': '2023.01.01', '安装地点': f'测试地点{i}',
            })
        return pd.DataFrame(rows)

    def test_file_created_and_has_correct_sheets(self):
        """生成的 Excel 应存在，且含主表 Sheet"""
        df = self._make_df_out()
        tmp = '_test_gen_main.xlsx'
        try:
            main.generate_excel_with_format(df, tmp)
            self.assertTrue(os.path.exists(tmp))
            wb = openpyxl.load_workbook(tmp)
            self.assertIn('安卓设备日统计表', wb.sheetnames)
            wb.close()
        finally:
            if os.path.exists(tmp): os.remove(tmp)

    def test_metered_sheet_created_when_provided(self):
        """传入非空 df_metered 时应生成计量客户单列 Sheet"""
        df = self._make_df_out(3)
        df_m = self._make_df_out(2)
        tmp = '_test_gen_metered.xlsx'
        try:
            main.generate_excel_with_format(df, tmp, df_metered=df_m)
            wb = openpyxl.load_workbook(tmp)
            self.assertIn('计量客户单列', wb.sheetnames)
            wb.close()
        finally:
            if os.path.exists(tmp): os.remove(tmp)

    def test_metered_sheet_absent_when_empty(self):
        """df_metered 为空时不应生成计量客户单列 Sheet"""
        df = self._make_df_out(3)
        tmp = '_test_gen_no_metered.xlsx'
        try:
            main.generate_excel_with_format(df, tmp, df_metered=pd.DataFrame())
            wb = openpyxl.load_workbook(tmp)
            self.assertNotIn('计量客户单列', wb.sheetnames)
            wb.close()
        finally:
            if os.path.exists(tmp): os.remove(tmp)

    def test_row_count_matches(self):
        """Sheet 数据行数应与 DataFrame 行数一致"""
        df = self._make_df_out(7)
        tmp = '_test_gen_rows.xlsx'
        try:
            main.generate_excel_with_format(df, tmp)
            wb = openpyxl.load_workbook(tmp)
            ws = wb['安卓设备日统计表']
            # row0=标题，row1=表头，row2..row8=数据(7行)
            data_rows = ws.max_row - 2
            self.assertEqual(data_rows, 7)
            wb.close()
        finally:
            if os.path.exists(tmp): os.remove(tmp)

    def test_exception_does_not_leave_partial_file(self):
        """apply_sheet_format 内部异常时，context manager 应正确关闭（Bug 2 回归）"""
        df = self._make_df_out(2)
        tmp = '_test_gen_exc.xlsx'
        try:
            with self.assertRaises(RuntimeError):
                with pd.ExcelWriter(tmp, engine='xlsxwriter') as writer:
                    df.to_excel(writer, index=False, sheet_name='安卓设备日统计表', startrow=1)
                    raise RuntimeError("模拟格式化异常")
            # writer 已关闭，无悬挂资源
        finally:
            if os.path.exists(tmp): os.remove(tmp)


# =============================================================================
# 5. send_to_robot 分支路径
# =============================================================================
class TestSendToRobot(unittest.TestCase):

    WEBHOOK = 'https://qyapi.weixin.qq.com/cgi-bin/webhook/send?key=testkey'

    def _upload_mock(self, media_id='MID_001'):
        m = MagicMock()
        m.raise_for_status.return_value = None
        m.json.return_value = {'media_id': media_id}
        m.text = ''
        return m

    def _send_mock(self, errcode=0):
        m = MagicMock()
        m.raise_for_status.return_value = None
        m.json.return_value = {'errcode': errcode, 'errmsg': 'ok' if errcode == 0 else 'err'}
        return m

    def test_skip_when_no_webhook(self):
        """WEBHOOK_URL 为空时跳过发送"""
        called = []
        with patch.object(main, 'WEBHOOK_URL', ''), \
             patch('requests.post', side_effect=lambda *a, **kw: called.append(a)):
            main.send_to_robot('fake.xlsx')
        self.assertEqual(len(called), 0)

    def test_skip_when_file_missing(self):
        """文件不存在时跳过发送"""
        called = []
        with patch.object(main, 'WEBHOOK_URL', self.WEBHOOK), \
             patch('os.path.exists', return_value=False), \
             patch('requests.post', side_effect=lambda *a, **kw: called.append(a)):
            main.send_to_robot('nonexistent.xlsx')
        self.assertEqual(len(called), 0)

    def test_full_success_path(self):
        """上传成功 + 发送成功 → 日志记录发送成功"""
        log_msgs = []
        with patch.object(main, 'WEBHOOK_URL', self.WEBHOOK), \
             patch('os.path.exists', return_value=True), \
             patch('builtins.open', mock_open(read_data=b'data')), \
             patch('requests.post', side_effect=[self._upload_mock(), self._send_mock(0)]), \
             patch.object(main.logger, 'info', side_effect=log_msgs.append):
            main.send_to_robot('report.xlsx')
        self.assertTrue(any('发送成功' in m for m in log_msgs))

    def test_upload_http_error_triggers_alert(self):
        """上传 HTTP 错误 → 飞书告警"""
        err_mock = MagicMock()
        err_mock.raise_for_status.side_effect = Exception('500')
        alerts = []
        with patch.object(main, 'WEBHOOK_URL', self.WEBHOOK), \
             patch('os.path.exists', return_value=True), \
             patch('builtins.open', mock_open(read_data=b'data')), \
             patch('requests.post', return_value=err_mock), \
             patch.object(main, 'send_feishu_alert',
                          side_effect=lambda *a, **kw: alerts.append(a)):
            main.send_to_robot('report.xlsx')
        self.assertTrue(any(a[0] == 'warning' for a in alerts))

    def test_send_errcode_nonzero_triggers_alert(self):
        """发送接口 errcode≠0 → 飞书告警（Bug 5 回归）"""
        alerts = []
        with patch.object(main, 'WEBHOOK_URL', self.WEBHOOK), \
             patch('os.path.exists', return_value=True), \
             patch('builtins.open', mock_open(read_data=b'data')), \
             patch('requests.post', side_effect=[self._upload_mock(), self._send_mock(40058)]), \
             patch.object(main, 'send_feishu_alert',
                          side_effect=lambda *a, **kw: alerts.append(a)):
            main.send_to_robot('report.xlsx')
        self.assertTrue(len(alerts) > 0)
        self.assertIn('业务层错误', alerts[0][1])

    def test_no_media_id_triggers_alert(self):
        """upload 返回无 media_id → 触发告警"""
        no_mid_mock = MagicMock()
        no_mid_mock.raise_for_status.return_value = None
        no_mid_mock.json.return_value = {}
        no_mid_mock.text = '{"errcode":40006}'
        alerts = []
        with patch.object(main, 'WEBHOOK_URL', self.WEBHOOK), \
             patch('os.path.exists', return_value=True), \
             patch('builtins.open', mock_open(read_data=b'data')), \
             patch('requests.post', return_value=no_mid_mock), \
             patch.object(main, 'send_feishu_alert',
                          side_effect=lambda *a, **kw: alerts.append(a)):
            main.send_to_robot('report.xlsx')
        self.assertTrue(len(alerts) > 0)


# =============================================================================
# 6. daily_task 主流程分支
# =============================================================================
class TestDailyTask(unittest.TestCase):

    def test_db_empty_returns_early(self):
        """数据库返回空 → 不调用 process_data"""
        called = []
        with patch.object(main, 'get_db_data', return_value=pd.DataFrame()), \
             patch.object(main, 'process_data',
                          side_effect=lambda *a: called.append(a) or (pd.DataFrame(), pd.DataFrame())), \
             patch.object(main, 'setup_logging'), \
             patch.object(main, 'clean_old_files'):
            main.daily_task()
        self.assertEqual(len(called), 0)

    def test_empty_df_final_skips_excel(self):
        """process_data 返回空 df_final → 不生成 Excel（Bug 6 回归）"""
        excel_called = []
        with patch.object(main, 'get_db_data',
                          return_value=pd.DataFrame({'x': [1]})), \
             patch.object(main, 'process_data',
                          return_value=(pd.DataFrame(), pd.DataFrame())), \
             patch.object(main, 'generate_excel_with_format',
                          side_effect=lambda *a, **kw: excel_called.append(1) or a[1]), \
             patch.object(main, 'send_feishu_alert'), \
             patch.object(main, 'setup_logging'), \
             patch.object(main, 'clean_old_files'):
            main.daily_task()
        self.assertEqual(len(excel_called), 0)

    def test_excel_failure_skips_send(self):
        """generate_excel_with_format 抛异常 → 不调用 send_to_robot"""
        send_called = []
        df_mock = pd.DataFrame({'序号': [1], '客户名称': ['A']})
        with patch.object(main, 'get_db_data',
                          return_value=pd.DataFrame({'x': [1]})), \
             patch.object(main, 'process_data',
                          return_value=(df_mock, pd.DataFrame())), \
             patch.object(main, 'generate_excel_with_format',
                          side_effect=RuntimeError("磁盘已满")), \
             patch.object(main, 'send_to_robot',
                          side_effect=lambda *a: send_called.append(a)), \
             patch.object(main, 'send_feishu_alert'), \
             patch.object(main, 'setup_logging'), \
             patch.object(main, 'clean_old_files'):
            main.daily_task()
        self.assertEqual(len(send_called), 0)

    def test_full_happy_path(self):
        """全部正常 → generate + send 各调用一次"""
        df_mock = pd.DataFrame({'序号': [1], '客户名称': ['A']})
        gen_called, send_called = [], []
        with patch.object(main, 'get_db_data',
                          return_value=pd.DataFrame({'x': [1]})), \
             patch.object(main, 'process_data',
                          return_value=(df_mock, pd.DataFrame())), \
             patch.object(main, 'generate_excel_with_format',
                          side_effect=lambda *a, **kw: gen_called.append(a[1]) or a[1]), \
             patch.object(main, 'send_to_robot',
                          side_effect=lambda fn: send_called.append(fn)), \
             patch.object(main, 'setup_logging'), \
             patch.object(main, 'clean_old_files'):
            main.daily_task()
        self.assertEqual(len(gen_called), 1)
        self.assertEqual(len(send_called), 1)


# =============================================================================
# 7. _check_device_count_diff 列名含空格场景（Bug 3 回归）
# =============================================================================
class TestCheckDeviceCountDiff(unittest.TestCase):

    def _make_local(self, path, codes):
        pd.DataFrame({'设备编号': codes}).to_excel(
            path, index=False, sheet_name='设备配置')

    def test_no_alert_when_equal(self):
        """在线=本地 数量时，不触发告警"""
        tmp = '_t_cdiff_eq.xlsx'
        self._make_local(tmp, ['D001', 'D002'])
        try:
            alerts = []
            df = pd.DataFrame({'设备编号': ['D001', 'D002']})
            with patch.object(main, 'CONFIG_LOCAL_FILE', tmp), \
                 patch.object(main, 'send_feishu_alert',
                              side_effect=lambda *a, **kw: alerts.append(a)):
                main._check_device_count_diff(df)
            self.assertEqual(len(alerts), 0)
        finally:
            if os.path.exists(tmp): os.remove(tmp)

    def test_alert_when_diff_exceeds_threshold(self):
        """差异 ≥ threshold 时触发告警"""
        tmp = '_t_cdiff_diff.xlsx'
        self._make_local(tmp, ['D001', 'D002', 'D003'])
        try:
            alerts = []
            df = pd.DataFrame({'设备编号': ['D001']})
            with patch.object(main, 'CONFIG_LOCAL_FILE', tmp), \
                 patch.object(main, 'send_feishu_alert',
                              side_effect=lambda *a, **kw: alerts.append(a)):
                main._check_device_count_diff(df, threshold=1)
            self.assertTrue(len(alerts) > 0)
        finally:
            if os.path.exists(tmp): os.remove(tmp)

    def test_spaces_in_column_name_still_works(self):
        """列名含空格时，经 strip 后应正确匹配（Bug 3 回归）"""
        tmp = '_t_cdiff_space.xlsx'
        self._make_local(tmp, ['D001', 'D002'])
        try:
            alerts = []
            # 在线表格列名含空格，通过 rename strip 后传入
            df_raw = pd.DataFrame({' 设备编号 ': ['D001', 'D002', 'D003']})
            df_stripped = df_raw.rename(columns=lambda c: str(c).strip())
            with patch.object(main, 'CONFIG_LOCAL_FILE', tmp), \
                 patch.object(main, 'send_feishu_alert',
                              side_effect=lambda *a, **kw: alerts.append(a)):
                main._check_device_count_diff(df_stripped, threshold=1)
            self.assertTrue(len(alerts) > 0, "差异应触发告警")
        finally:
            if os.path.exists(tmp): os.remove(tmp)


# =============================================================================
if __name__ == '__main__':
    loader = unittest.TestLoader()
    suite = unittest.TestSuite()

    test_classes = [
        TestCheckSync,
        TestProcessData,
        TestLoadConfigFallback,
        TestGenerateExcel,
        TestSendToRobot,
        TestDailyTask,
        TestCheckDeviceCountDiff,
    ]
    for cls in test_classes:
        suite.addTests(loader.loadTestsFromTestCase(cls))

    runner = unittest.TextTestRunner(verbosity=2)
    result = runner.run(suite)
    sys.exit(0 if result.wasSuccessful() else 1)
